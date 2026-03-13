"""
Excel workbook builder for the daily market report.

The workbook keeps rolling history and exposes dedicated sheets for:
- current report summary
- CO2 history and 7/30 day views
- yearly BASE energy snapshots
- electricity spot hourly history from PSE
- gas history and 7/30 day views
"""
from __future__ import annotations

import logging
from datetime import timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

SHEETS = {
    "report": "Raport_dzienny",
    "co2_history": "CO2_historia",
    "co2_7d": "CO2_7D",
    "co2_30d": "CO2_30D",
    "energy_base": "Energia_BASE_hist",
    "power_spot": "Spot_energia_hist",
    "power_spot_day": "Spot_energia_dzien",
    "power_spot_7d": "Spot_energia_7D",
    "power_spot_30d": "Spot_energia_30D",
    "gas_history": "Gaz_historia",
    "gas_7d": "Gaz_7D",
    "gas_30d": "Gaz_30D",
}

POWER_SPOT_COLUMNS = [
    "Data_Dostawy",
    "Godzina_Od",
    "Godzina_Do",
    "Godzina_Label",
    "Cena_SPOT_PLN_MWh",
    "Cena_Min_Kwadrans_PLN_MWh",
    "Cena_Max_Kwadrans_PLN_MWh",
    "Liczba_Kwadransow",
    "Data_Publikacji",
    "Data_Pobrania",
    "Zrodlo_URL",
    "Interwal_Zrodla",
]

ENERGY_BASE_COLUMNS = [
    "Data_Raportu",
    "Data_Pobrania",
    "Zrodlo_URL",
    "Kontrakt_2026",
    "Cena_BASE_2026_PLN_MWh",
    "Data_Notowania_2026",
    "Status_2026",
    "Kontrakt_2027",
    "Cena_BASE_2027_PLN_MWh",
    "Data_Notowania_2027",
    "Status_2027",
    "Kontrakt_2028",
    "Cena_BASE_2028_PLN_MWh",
    "Data_Notowania_2028",
    "Status_2028",
    "Kolumna_Ceny_2026",
    "URL_Notowania_2026",
    "Kolumna_Ceny_2027",
    "URL_Notowania_2027",
    "Kolumna_Ceny_2028",
    "URL_Notowania_2028",
]

GAS_COLUMNS = [
    "Data_Raportu",
    "Data_Pobrania",
    "Data_Notowania",
    "Indeks",
    "Cena_Biezaca_PLN_MWh",
    "Zmiana_Proc",
    "Wolumen_MWh",
    "Kolumna_Zrodla_Ceny",
    "Zrodlo_URL",
]


def _apply_header_style(ws) -> None:
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"


def _auto_column_width(ws) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 50))


def _resolve_output_path(config: dict) -> Path:
    data_cfg = config.get("data", {})
    filename = data_cfg.get("output_file", "tge_dane_historyczne.xlsx")
    output_dir = data_cfg.get("output_dir", "") or ""
    if output_dir:
        path = Path(output_dir) / filename
        path.parent.mkdir(parents=True, exist_ok=True)
    else:
        path = Path(filename)
    return path


def _read_existing_sheets(path: Path) -> dict[str, pd.DataFrame]:
    if not path.exists():
        return {}

    sheets: dict[str, pd.DataFrame] = {}
    try:
        xf = pd.ExcelFile(path, engine="openpyxl")
        for sheet in xf.sheet_names:
            sheets[sheet] = xf.parse(sheet)
        xf.close()
    except Exception as exc:
        logger.warning("Could not read existing workbook %s: %s", path, exc)
    return sheets


def _to_datetime_series(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series([pd.NaT] * len(df), index=df.index)

    raw = df[column]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        series = pd.to_datetime(raw, format=fmt, errors="coerce")
        if not series.isna().all():
            return series

    series = pd.to_datetime(raw, errors="coerce", dayfirst=True)
    if series.isna().all():
        series = pd.to_datetime(raw, errors="coerce")
    return series


def _numeric_series(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series(dtype="float64")
    return pd.to_numeric(df[column], errors="coerce")


def _sort_if_possible(df: pd.DataFrame, columns: list[str], ascending: bool = False) -> pd.DataFrame:
    existing_columns = [column for column in columns if column in df.columns]
    if not existing_columns:
        return df.copy()
    return df.sort_values(by=existing_columns, ascending=ascending)


def _sort_spot_history(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    result = df.copy()
    result["__day"] = _to_datetime_series(result, "Data_Dostawy")
    result["__hour"] = _to_datetime_series(result, "Godzina_Od")
    result = result.sort_values(["__day", "__hour"], ascending=[False, True])
    return result.drop(columns=["__day", "__hour"]).reset_index(drop=True)


def _normalize_power_spot_history(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=POWER_SPOT_COLUMNS)

    result = df.copy()
    for column in POWER_SPOT_COLUMNS:
        if column not in result.columns:
            result[column] = pd.NA

    result = result[POWER_SPOT_COLUMNS]
    if "Data_Dostawy" in result.columns:
        result = result[result["Data_Dostawy"].notna()].copy()
    return result.reset_index(drop=True)


def _normalize_energy_history(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=ENERGY_BASE_COLUMNS)

    result = df.copy()
    for column in ENERGY_BASE_COLUMNS:
        if column not in result.columns:
            result[column] = pd.NA

    result = result[ENERGY_BASE_COLUMNS]
    result = result[result["Data_Raportu"].notna()].copy()
    return result.reset_index(drop=True)


def _normalize_gas_history(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=GAS_COLUMNS)

    result = df.copy()
    for column in GAS_COLUMNS:
        if column not in result.columns:
            result[column] = pd.NA

    result = result[GAS_COLUMNS]
    result = result[result["Data_Raportu"].notna()].copy()
    return result.reset_index(drop=True)


def _merge_history(
    existing: pd.DataFrame | None,
    incoming: pd.DataFrame | None,
    key_columns: list[str],
    sort_columns: list[str],
) -> pd.DataFrame:
    frames = []
    if existing is not None and not existing.empty:
        frames.append(existing.copy())
    if incoming is not None and not incoming.empty:
        frames.append(incoming.copy())

    if not frames:
        return pd.DataFrame()

    merged = pd.concat(frames, ignore_index=True, sort=False)

    sort_helpers: list[str] = []
    for column in sort_columns:
        helper = f"__sort_{column}"
        merged[helper] = _to_datetime_series(merged, column)
        sort_helpers.append(helper)

    merged = merged.sort_values(sort_helpers)
    existing_keys = [column for column in key_columns if column in merged.columns]
    if existing_keys:
        merged = merged.drop_duplicates(subset=existing_keys, keep="last")
    merged = merged.drop(columns=sort_helpers)
    merged = merged.reset_index(drop=True)
    return merged


def _last_days(df: pd.DataFrame, date_column: str, days: int) -> pd.DataFrame:
    if df.empty or date_column not in df.columns:
        return df.copy()

    subset = df.copy()
    subset["__date"] = _to_datetime_series(subset, date_column)
    subset = subset.dropna(subset=["__date"])
    if subset.empty:
        return df.iloc[0:0].copy()

    latest_date = subset["__date"].max().normalize()
    cutoff = latest_date - timedelta(days=days - 1)
    subset = subset[subset["__date"] >= cutoff].copy()
    subset = subset.sort_values("__date", ascending=False).drop(columns=["__date"])
    return subset.reset_index(drop=True)


def _latest_row(df: pd.DataFrame, date_column: str) -> pd.Series | None:
    if df.empty or date_column not in df.columns:
        return None

    subset = df.copy()
    subset["__date"] = _to_datetime_series(subset, date_column)
    subset = subset.dropna(subset=["__date"])
    subset = subset.sort_values("__date")
    if subset.empty:
        return None
    return subset.iloc[-1]


def _format_value(value: object, decimals: int = 2) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "brak"
    if isinstance(value, (int, float)):
        return f"{value:.{decimals}f}"
    return str(value)


def _build_power_spot_daily_summary(power_spot_history: pd.DataFrame) -> pd.DataFrame:
    if power_spot_history.empty:
        return pd.DataFrame(
            columns=[
                "Data_Dostawy",
                "Cena_Srednia_Dzien_PLN_MWh",
                "Cena_Min_Dzien_PLN_MWh",
                "Cena_Max_Dzien_PLN_MWh",
                "Liczba_Godzin",
            ]
        )

    daily = power_spot_history.copy()
    daily["Cena_SPOT_PLN_MWh"] = pd.to_numeric(daily["Cena_SPOT_PLN_MWh"], errors="coerce")
    daily = daily.dropna(subset=["Cena_SPOT_PLN_MWh"])
    if daily.empty:
        return pd.DataFrame()

    grouped = (
        daily.groupby("Data_Dostawy", as_index=False)
        .agg(
            Cena_Srednia_Dzien_PLN_MWh=("Cena_SPOT_PLN_MWh", "mean"),
            Cena_Min_Dzien_PLN_MWh=("Cena_SPOT_PLN_MWh", "min"),
            Cena_Max_Dzien_PLN_MWh=("Cena_SPOT_PLN_MWh", "max"),
            Liczba_Godzin=("Cena_SPOT_PLN_MWh", "count"),
        )
    )
    return grouped.sort_values("Data_Dostawy").reset_index(drop=True)


def _latest_spot_day_rows(power_spot_history: pd.DataFrame) -> pd.DataFrame:
    if power_spot_history.empty or "Data_Dostawy" not in power_spot_history.columns:
        return pd.DataFrame(columns=power_spot_history.columns if isinstance(power_spot_history, pd.DataFrame) else [])

    subset = power_spot_history.copy()
    subset["__day"] = _to_datetime_series(subset, "Data_Dostawy")
    subset = subset.dropna(subset=["__day"])
    if subset.empty:
        return power_spot_history.iloc[0:0].copy()

    latest_day = subset["__day"].max().strftime("%Y-%m-%d")
    subset = subset[subset["Data_Dostawy"].astype(str) == latest_day].copy()
    subset["__hour"] = _to_datetime_series(subset, "Godzina_Od")
    subset = subset.sort_values("__hour").drop(columns=["__day", "__hour"])
    return subset.reset_index(drop=True)


def _build_range_rows(
    section: str,
    history: pd.DataFrame,
    value_column: str,
    date_column: str,
    unit: str,
    current_label: str = "Cena biezaca",
    current_note: str = "Najnowsza dostepna wartosc.",
    sheet_7d: str | None = None,
    sheet_30d: str | None = None,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    latest = _latest_row(history, date_column)

    if latest is None:
        rows.append(
            {
                "Sekcja": section,
                "Metryka": "Status",
                "Wartosc": "brak danych",
                "Jednostka": unit,
                "Data": "",
                "Uwagi": "Arkusz zostanie uzupelniony po pierwszym poprawnym pobraniu.",
            }
        )
        return rows

    rows.append(
        {
            "Sekcja": section,
            "Metryka": current_label,
            "Wartosc": latest.get(value_column),
            "Jednostka": unit,
            "Data": latest.get(date_column, ""),
            "Uwagi": current_note,
        }
    )

    for window, sheet_name in ((7, sheet_7d), (30, sheet_30d)):
        scoped = _last_days(history, date_column, window)
        values = _numeric_series(scoped, value_column).dropna()
        note = f"Szczegoly w arkuszu {sheet_name}" if sheet_name else ""

        if values.empty:
            rows.append(
                {
                    "Sekcja": section,
                    "Metryka": f"Zakres {window}D",
                    "Wartosc": "brak danych",
                    "Jednostka": unit,
                    "Data": latest.get(date_column, ""),
                    "Uwagi": "Historia jest jeszcze zbyt krotka.",
                }
            )
            continue

        rows.extend(
            [
                {
                    "Sekcja": section,
                    "Metryka": f"Minimum {window}D",
                    "Wartosc": values.min(),
                    "Jednostka": unit,
                    "Data": latest.get(date_column, ""),
                    "Uwagi": note,
                },
                {
                    "Sekcja": section,
                    "Metryka": f"Maksimum {window}D",
                    "Wartosc": values.max(),
                    "Jednostka": unit,
                    "Data": latest.get(date_column, ""),
                    "Uwagi": note,
                },
                {
                    "Sekcja": section,
                    "Metryka": f"Srednia {window}D",
                    "Wartosc": values.mean(),
                    "Jednostka": unit,
                    "Data": latest.get(date_column, ""),
                    "Uwagi": f"Liczba obserwacji: {len(values)}",
                },
            ]
        )

    return rows


def _build_spot_rows(power_spot_history: pd.DataFrame) -> list[dict[str, object]]:
    daily_summary = _build_power_spot_daily_summary(power_spot_history)
    latest_day = _latest_row(daily_summary, "Data_Dostawy")
    if latest_day is None:
        return [
            {
                "Sekcja": "Spot energia",
                "Metryka": "Status",
                "Wartosc": "brak danych",
                "Jednostka": "PLN/MWh",
                "Data": "",
                "Uwagi": "Brak danych godzinowych z PSE.",
            }
        ]

    rows = [
        {
            "Sekcja": "Spot energia",
            "Metryka": "Srednia dnia",
            "Wartosc": latest_day.get("Cena_Srednia_Dzien_PLN_MWh"),
            "Jednostka": "PLN/MWh",
            "Data": latest_day.get("Data_Dostawy", ""),
            "Uwagi": f"{int(latest_day.get('Liczba_Godzin', 0))} godzin w arkuszu {SHEETS['power_spot_day']}",
        },
        {
            "Sekcja": "Spot energia",
            "Metryka": "Minimum dnia",
            "Wartosc": latest_day.get("Cena_Min_Dzien_PLN_MWh"),
            "Jednostka": "PLN/MWh",
            "Data": latest_day.get("Data_Dostawy", ""),
            "Uwagi": f"Szczegoly godzinowe w arkuszu {SHEETS['power_spot_day']}",
        },
        {
            "Sekcja": "Spot energia",
            "Metryka": "Maksimum dnia",
            "Wartosc": latest_day.get("Cena_Max_Dzien_PLN_MWh"),
            "Jednostka": "PLN/MWh",
            "Data": latest_day.get("Data_Dostawy", ""),
            "Uwagi": f"Szczegoly godzinowe w arkuszu {SHEETS['power_spot_day']}",
        },
    ]
    rows.extend(
        _build_range_rows(
            section="Spot energia",
            history=daily_summary,
            value_column="Cena_Srednia_Dzien_PLN_MWh",
            date_column="Data_Dostawy",
            unit="PLN/MWh",
            current_label="Srednia dnia",
            current_note=f"{int(latest_day.get('Liczba_Godzin', 0))} godzin w arkuszu {SHEETS['power_spot_day']}",
            sheet_7d=SHEETS["power_spot_7d"],
            sheet_30d=SHEETS["power_spot_30d"],
        )[1:]
    )
    return rows


def _build_report_sheet(
    co2_history: pd.DataFrame,
    energy_history: pd.DataFrame,
    power_spot_history: pd.DataFrame,
    gas_history: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []

    rows.extend(
        _build_range_rows(
            section="CO2",
            history=co2_history,
            value_column="Cena_CO2",
            date_column="Data",
            unit="EUR",
            sheet_7d=SHEETS["co2_7d"],
            sheet_30d=SHEETS["co2_30d"],
        )
    )

    latest_energy = _latest_row(energy_history, "Data_Raportu")
    if latest_energy is None:
        rows.append(
            {
                "Sekcja": "Energia BASE",
                "Metryka": "Status",
                "Wartosc": "brak danych",
                "Jednostka": "PLN/MWh",
                "Data": "",
                "Uwagi": "Brak danych o kontraktach BASE.",
            }
        )
    else:
        for year in (2026, 2027, 2028):
            rows.append(
                {
                    "Sekcja": "Energia BASE",
                    "Metryka": str(year),
                    "Wartosc": latest_energy.get(f"Cena_BASE_{year}_PLN_MWh"),
                    "Jednostka": "PLN/MWh",
                    "Data": latest_energy.get(f"Data_Notowania_{year}") or latest_energy.get("Data_Raportu", ""),
                    "Uwagi": latest_energy.get(f"Status_{year}", ""),
                }
            )

    rows.extend(_build_spot_rows(power_spot_history))

    rows.extend(
        _build_range_rows(
            section="Gaz",
            history=gas_history,
            value_column="Cena_Biezaca_PLN_MWh",
            date_column="Data_Raportu",
            unit="PLN/MWh",
            sheet_7d=SHEETS["gas_7d"],
            sheet_30d=SHEETS["gas_30d"],
        )
    )

    report = pd.DataFrame(rows)
    return report[["Sekcja", "Metryka", "Wartosc", "Jednostka", "Data", "Uwagi"]]


def append_to_excel(scraped: dict[str, pd.DataFrame], config: dict) -> Path:
    output_path = _resolve_output_path(config)
    existing = _read_existing_sheets(output_path)

    co2_history = _merge_history(
        existing.get(SHEETS["co2_history"]),
        scraped.get("co2_history"),
        key_columns=["Data"],
        sort_columns=["Data", "Data_Pobrania"],
    )
    energy_history = _merge_history(
        _normalize_energy_history(existing.get(SHEETS["energy_base"])),
        _normalize_energy_history(scraped.get("energy_base_history")),
        key_columns=["Data_Raportu"],
        sort_columns=["Data_Raportu", "Data_Pobrania"],
    )
    energy_history = _normalize_energy_history(energy_history)
    power_spot_history = _merge_history(
        _normalize_power_spot_history(existing.get(SHEETS["power_spot"])),
        _normalize_power_spot_history(scraped.get("power_spot_history")),
        key_columns=["Data_Dostawy", "Godzina_Od", "Godzina_Do"],
        sort_columns=["Data_Dostawy", "Godzina_Od", "Data_Pobrania"],
    )
    power_spot_history = _normalize_power_spot_history(power_spot_history)
    gas_history = _merge_history(
        _normalize_gas_history(existing.get(SHEETS["gas_history"])),
        _normalize_gas_history(scraped.get("gas_spot_history")),
        key_columns=["Data_Raportu", "Indeks"],
        sort_columns=["Data_Raportu", "Data_Pobrania"],
    )
    gas_history = _normalize_gas_history(gas_history)

    report = _build_report_sheet(co2_history, energy_history, power_spot_history, gas_history)

    sheets = {
        SHEETS["report"]: report,
        SHEETS["co2_history"]: _sort_if_possible(co2_history, ["Data"], ascending=False),
        SHEETS["co2_7d"]: _last_days(co2_history, "Data", 7),
        SHEETS["co2_30d"]: _last_days(co2_history, "Data", 30),
        SHEETS["energy_base"]: _sort_if_possible(energy_history, ["Data_Raportu"], ascending=False),
        SHEETS["power_spot"]: _sort_spot_history(power_spot_history),
        SHEETS["power_spot_day"]: _latest_spot_day_rows(power_spot_history),
        SHEETS["power_spot_7d"]: _sort_spot_history(_last_days(power_spot_history, "Data_Dostawy", 7)),
        SHEETS["power_spot_30d"]: _sort_spot_history(_last_days(power_spot_history, "Data_Dostawy", 30)),
        SHEETS["gas_history"]: _sort_if_possible(gas_history, ["Data_Raportu"], ascending=False),
        SHEETS["gas_7d"]: _last_days(gas_history, "Data_Raportu", 7),
        SHEETS["gas_30d"]: _last_days(gas_history, "Data_Raportu", 30),
    }

    _write_excel(output_path, sheets)
    logger.info("Saved workbook with %d sheets to %s", len(sheets), output_path)
    return output_path


def _write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            frame = df if isinstance(df, pd.DataFrame) else pd.DataFrame()
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    try:
        wb = load_workbook(str(path))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row >= 1:
                _apply_header_style(ws)
                _auto_column_width(ws)
        wb.save(str(path))
    except Exception as exc:
        logger.warning("Excel formatting failed for %s: %s", path, exc)


def get_summary(output_path: Path) -> str:
    if not output_path.exists():
        return "Brak pliku z danymi."

    try:
        sheets = _read_existing_sheets(output_path)
        report = sheets.get(SHEETS["report"], pd.DataFrame())
        co2_7d = sheets.get(SHEETS["co2_7d"], pd.DataFrame())
        co2_30d = sheets.get(SHEETS["co2_30d"], pd.DataFrame())
        spot_day = sheets.get(SHEETS["power_spot_day"], pd.DataFrame())
        spot_30d = sheets.get(SHEETS["power_spot_30d"], pd.DataFrame())
        gas_7d = sheets.get(SHEETS["gas_7d"], pd.DataFrame())
        gas_30d = sheets.get(SHEETS["gas_30d"], pd.DataFrame())

        lines = [f"Plik: {output_path.name}"]
        lines.append("Arkusze: " + ", ".join(SHEETS.values()))

        if not report.empty:
            energy_rows = report[report["Sekcja"] == "Energia BASE"]
            co2_current = report[(report["Sekcja"] == "CO2") & (report["Metryka"] == "Cena biezaca")]
            spot_current = report[(report["Sekcja"] == "Spot energia") & (report["Metryka"] == "Srednia dnia")]
            gas_current = report[(report["Sekcja"] == "Gaz") & (report["Metryka"] == "Cena biezaca")]

            if not co2_current.empty:
                row = co2_current.iloc[0]
                lines.append(f"CO2: {_format_value(row['Wartosc'])} {row['Jednostka']} ({row['Data']})")

            if not energy_rows.empty:
                energy_parts = []
                for _, row in energy_rows.iterrows():
                    energy_parts.append(f"{row['Metryka']}: {_format_value(row['Wartosc'])}")
                lines.append("Energia BASE: " + ", ".join(energy_parts))

            if not spot_current.empty:
                row = spot_current.iloc[0]
                lines.append(
                    f"Spot energia: {_format_value(row['Wartosc'])} {row['Jednostka']} ({row['Data']}, godzin: {len(spot_day)})"
                )

            if not gas_current.empty:
                row = gas_current.iloc[0]
                lines.append(f"Gaz: {_format_value(row['Wartosc'])} {row['Jednostka']} ({row['Data']})")

        lines.append(f"CO2 7D: {len(co2_7d)} rekordow | CO2 30D: {len(co2_30d)} rekordow")
        lines.append(f"SPOT 30D: {len(spot_30d)} rekordow godzinowych | SPOT dzien: {len(spot_day)} rekordow")
        lines.append(f"Gaz 7D: {len(gas_7d)} rekordow | Gaz 30D: {len(gas_30d)} rekordow")
        return "\n".join(lines)
    except Exception as exc:
        return f"Blad odczytu raportu: {exc}"
